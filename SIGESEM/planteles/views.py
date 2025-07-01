from django.shortcuts import render, redirect
from .models import *
from .forms import *
from django.contrib import messages
from django.contrib.auth.models import User, Group
from django.shortcuts import redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse
from usuarios.models import PerfilUsuario 
import openpyxl
from openpyxl import Workbook
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from urllib.parse import quote
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


#Panel principal
@login_required
def dashboard(request):
    total_discentes = Discente.objects.count()
    total_bajas = Baja.objects.count()
    total_casos = CasoMedicoLegal.objects.count()
    total_planteles = Plantel.objects.count()

    return render(request, 'planteles/dashboard.html', {
        'total_discentes': total_discentes,
        'total_bajas': total_bajas,
        'total_casos': total_casos,
        'total_planteles': total_planteles,
    })



@login_required
def dashboard(request):
    total_discentes = Discente.objects.count()
    total_hombres = Discente.objects.filter(genero='M').count()
    total_mujeres = Discente.objects.filter(genero='F').count()
    total_bajas = Baja.objects.count()
    total_casos = CasoMedicoLegal.objects.count()
    total_planteles = Plantel.objects.count()

    contexto = {
        'total_discentes': total_discentes,
        'total_hombres': total_hombres,
        'total_mujeres': total_mujeres,
        'total_bajas': total_bajas,
        'total_casos': total_casos,
        'total_planteles': total_planteles
    }
    return render(request, 'planteles/dashboard.html', contexto)


#Discentes

def crear_objeto(request, modelo_form, template='planteles/agregar_discente.html'):
    form = modelo_form(request.POST or None)
    if form.is_valid():
        form.save()
        return redirect('lista')
    return render(request, template, {'form': form})




@login_required
def lista_discentes(request):
    genero = request.GET.get('genero')
    if genero in ['M', 'F', 'O']:
        discentes = Discente.objects.filter(genero=genero)
    else:
        discentes = Discente.objects.all()

    perfil = request.user.perfilusuario  # accede al perfil
    puede_crud = perfil.es_encargado  # solo encargado puede modificar

    contexto = {
        'objetos': discentes,
        'titulo': 'Discentes',
        'puede_crud': puede_crud,
    }
    return render(request, 'planteles/lista.html', contexto)




def editar_discente(request, pk):
    discente = get_object_or_404(Discente, pk=pk)
    if request.method == 'POST':
        form = DiscenteForm(request.POST, instance=discente)
        if form.is_valid():
            form.save()
            messages.success(request, 'Discente actualizado correctamente.')
            return redirect('lista')
    else:
        form = DiscenteForm(instance=discente)

    return render(request, 'planteles/formulario.html', {
        'form': form,
        'titulo': 'Editar Discente'
    })


#baja Discente

def nuevo_tipo_baja(request):
    if request.method == 'POST':
        form = TipoBajaForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tipo de baja registrado correctamente.')
            return redirect('nuevo_tipo_baja')
    else:
        form = TipoBajaForm()
    return render(request, 'planteles/form_tipo_baja.html', {'form': form})




def nueva_baja(request):
    if request.method == 'POST':
        form = BajaForm(request.POST)
        if form.is_valid():
            matricula = form.cleaned_data['matricula'].strip().upper()
            try:
                discente = Discente.objects.get(matricula=matricula)

                # Guardar la baja
                nueva_baja = Baja(
                    id_discente=discente,
                    tipo_baja=form.cleaned_data['tipo_baja'],
                    fecha_baja=form.cleaned_data['fecha_baja'],
                    motivo=form.cleaned_data['motivo']
                )
                nueva_baja.save()

                # Actualizar EfectivoBaja
                efectivo, creado = EfectivoBaja.objects.get_or_create(id_plantel=discente.id_plantel)
                if discente.genero == 'M':
                    efectivo.cant_hombre += 1
                elif discente.genero == 'F':
                    efectivo.cant_mujer += 1
                efectivo.cant_total = efectivo.cant_hombre + efectivo.cant_mujer
                efectivo.save()

                # Eliminar al discente
                discente.delete()

                messages.success(request, "✅ Baja registrada correctamente.")
                return redirect('nueva_baja')
            except Discente.DoesNotExist:
                messages.error(request, "❌ No se encontró un discente con esa matrícula.")
        else:
            messages.error(request, "❌ Corrige los errores del formulario.")
    else:
        form = BajaForm()

    return render(request, 'planteles/form_baja.html', {'form': form})

def editar_baja(request, id):
    baja = get_object_or_404(Baja, id=id)
    if request.method == 'POST':
        form = BajaForm(request.POST, instance=baja)
        if form.is_valid():
            form.save()
            messages.success(request, "Baja actualizada correctamente.")
            return redirect('lista_bajas')
    else:
        form = BajaForm(instance=baja)
    return render(request, 'planteles/formulario_baja.html', {
        'form': form,
        'titulo': 'Editar Baja'
    })

def eliminar_baja(request, id):
    baja = get_object_or_404(Baja, id=id)
    if request.method == 'POST':
        baja.delete()
        messages.success(request, "Baja eliminada correctamente.")
        return redirect('lista_bajas')
    # En caso de GET, podrías mostrar una confirmación o redirigir
    return render(request, 'planteles/confirmar_eliminacion.html', {
        'objeto': baja,
        'titulo': 'Confirmar eliminación de baja'
    })


@login_required
def tu_vista(request):
    ...




@login_required
def lista_bajas(request):
    bajas = Baja.objects.select_related('id_discente', 'tipo_baja').all().order_by('-fecha_baja')

    perfil = getattr(request.user, 'perfilusuario', None)
    puede_crud = perfil and (perfil.es_encargado)

    contexto = {
        'objetos': bajas,
        'titulo': 'Bajas registradas',
        'puede_crud': puede_crud,
    }
    return render(request, 'planteles/lista_bajas.html', contexto)



#Plantel

@login_required
def lista_planteles(request):
    user = request.user



    rol = user.perfilusuario.rol

    if rol == 'encargado':
        return redirect('dashboard')

    planteles = Plantel.objects.all()
    puede_crud = (rol == 'administrador')

    contexto = {
        'planteles': planteles,
        'puede_crud': puede_crud,
        'titulo': 'Lista de Planteles',
    }
    return render(request, 'planteles/lista_planteles.html', contexto)


def nuevo_plantel(request):
    if request.method == 'POST':
        form = PlantelForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_planteles')
    else:
        form = PlantelForm()
    return render(request, 'planteles/formulario_plantel.html', {'form': form, 'titulo': 'Nuevo Plantel'})

def editar_plantel(request, id):
    plantel = get_object_or_404(Plantel, id=id)
    if request.method == 'POST':
        form = PlantelForm(request.POST, instance=plantel)
        if form.is_valid():
            form.save()
            return redirect('lista_planteles')
    else:
        form = PlantelForm(instance=plantel)
    return render(request, 'planteles/formulario_plantel.html', {'form': form, 'titulo': 'Editar Plantel'})

def eliminar_plantel(request, id):
    plantel = get_object_or_404(Plantel, id=id)
    if request.method == 'POST':
        plantel.delete()
        return redirect('lista_planteles')
    return render(request, 'planteles/confirmar_eliminar.html', {'plantel': plantel})




def lista_usuarios(request):
    usuarios = User.objects.all()
    return render(request, 'usuarios/lista.html', {'usuarios': usuarios})

#Casos medicos legales

def nuevo_caso_medico_legal(request):
    if request.method == 'POST':
        form = CasoMedicoLegalForm(request.POST)
        if form.is_valid():
            caso = form.save(commit=False)
            caso.id_discente = form.cleaned_data['matricula']  # Aquí matricula es el objeto Discente
            caso.save()
            messages.success(request, 'El caso médico-legal se ha guardado exitosamente.')
            return redirect('nuevo_caso')
    else:
        form = CasoMedicoLegalForm()
    return render(request, 'casos/formulario.html', {
        'form': form,
        'titulo': 'Registrar Caso Médico-Legal'
    })


def editar_caso(request, id):
    caso = get_object_or_404(CasoMedicoLegal, id=id)
    if request.method == 'POST':
        form = CasoMedicoLegalForm(request.POST, instance=caso)
        if form.is_valid():
            caso = form.save(commit=False)
            caso.id_discente = form.cleaned_data['matricula']  # matricula es objeto Discente
            caso.save()
            messages.success(request, 'El caso médico-legal se ha actualizado exitosamente.')
            return redirect('lista_casos')
    else:
        form = CasoMedicoLegalForm(instance=caso)

    return render(request, 'casos/formulario.html', {
        'form': form,
        'titulo': 'Editar Caso Médico-Legal'
    })


def eliminar_caso(request, pk):
    caso = get_object_or_404(CasoMedicoLegal, pk=pk)
    caso.delete()
    messages.success(request, 'El caso médico-legal ha sido eliminado correctamente.')
    return redirect('lista_casos')

@login_required
def lista_casos(request):
    casos = CasoMedicoLegal.objects.select_related('id_discente', 'id_discente__id_plantel').all()

    perfil = getattr(request.user, 'perfilusuario', None)
    puede_crud = perfil and (perfil.es_encargado)

    contexto = {
        'titulo': 'Casos Médico-Legales Registrados',
        'objetos': casos,
        'puede_crud': puede_crud,
    }
    return render(request, 'planteles/lista_casos.html', contexto)


 # Exportar en excel


def exportar_discentes_excel(request):
    genero = request.GET.get('genero')
    if genero in ['M', 'F', 'O']:
        discentes = Discente.objects.select_related(
            'id_plantel', 'id_categoria', 'id_entidad'
        ).filter(genero=genero)
    else:
        discentes = Discente.objects.select_related(
            'id_plantel', 'id_categoria', 'id_entidad'
        ).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Discentes"

    # Encabezados
    headers = [
        "Matrícula", "Nombre completo", "Género", "Fecha de Nacimiento", "Plantel",
        "Categoría", "Entidad", "Fecha Ingreso", "Antigüedad"
    ]
    ws.append(headers)

    # Estilo de encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Agregar datos
    for d in discentes:
        row = [
            d.matricula,
            f"{d.nombre} {d.apellido}",
            d.get_genero_display(),
            d.fecha_nacimiento,
            d.id_plantel.nombre,
            d.id_categoria.nombre_categoria,
            d.id_entidad.nombre_entidad,
            d.fecha_ingreso,
            d.antiguedad
        ]
        ws.append(row)

    # Aplicar formato de fecha (col 4 y 8 -> índices 3 y 7 base cero)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col in [3, 7]:  # Columnas con fechas
            if row[col].value:
                row[col].number_format = 'DD/MM/YYYY'
                row[col].alignment = Alignment(horizontal="center")

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2

    # Preparar respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )

    nombre = f"discentes_{genero.lower()}.xlsx" if genero else "discentes.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{quote(nombre)}"'

    wb.save(response)
    return response


@login_required
def exportar_bajas_excel(request):
    bajas = Baja.objects.select_related('id_discente', 'tipo_baja', 'id_discente__id_plantel').all().order_by('-fecha_baja')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bajas Registradas"

    # Encabezados
    headers = ["#", "Matrícula", "Nombre completo", "Plantel", "Tipo de baja", "Fecha", "Motivo"]
    ws.append(headers)

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")  # Azul
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Aplicar estilos a encabezados
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=column_title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Rellenar filas con datos
    for idx, baja in enumerate(bajas, start=1):
        ws.append([
            idx,
            baja.id_discente.matricula,
            f"{baja.id_discente.nombre} {baja.id_discente.apellido}",
            baja.id_discente.id_plantel.nombre,
            baja.tipo_baja.nombre,
            baja.fecha_baja,
            baja.motivo
        ])

    # Aplicar formato a las filas
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            # Alinear texto a la izquierda para texto y centro para fechas o números
            if cell.column == 1 or cell.column == 6:  # # y Fecha
                cell.alignment = center_align
            else:
                cell.alignment = left_align
            # Formatear fecha
            if cell.column == 6:
                cell.number_format = 'DD/MM/YYYY'

    # Ajustar ancho de columnas
    column_widths = {
        1: 5,   # #
        2: 15,  # Matrícula
        3: 30,  # Nombre completo
        4: 25,  # Plantel
        5: 20,  # Tipo de baja
        6: 15,  # Fecha
        7: 40   # Motivo
    }
    for col_num, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width

    # Preparar la respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    filename = "bajas_registradas.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{quote(filename)}"'

    wb.save(response)
    return response



@login_required
def exportar_casos_excel(request):
    casos = CasoMedicoLegal.objects.select_related('id_discente', 'id_discente__id_plantel').all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Casos Médico-Legales"

    headers = ["#", "Matrícula", "Nombre completo", "Plantel", "Fecha del caso", "Descripción", "Acciones adoptadas"]
    ws.append(headers)

    # Estilos encabezado
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Datos
    for idx, caso in enumerate(casos, start=1):
        ws.append([
            idx,
            caso.id_discente.matricula,
            f"{caso.id_discente.nombre} {caso.id_discente.apellido}",
            caso.id_discente.id_plantel.nombre,
            caso.fecha_caso.strftime("%d/%m/%Y"),
            caso.descripcion,
            caso.acciones_adoptadas,
        ])

    # Ajustar ancho columnas
    column_widths = [5, 15, 30, 20, 15, 50, 50]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=casos_medico_legales.xlsx'
    wb.save(response)
    return response


@login_required
def exportar_casos_pdf(request):
    casos = CasoMedicoLegal.objects.select_related('id_discente', 'id_discente__id_plantel').all()

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter
    styles = getSampleStyleSheet()
    text_style = styles["Normal"]

    # Título
    c.setFont("Helvetica-Bold", 16)
    c.drawString(40, height - 50, "Reporte de Casos Médico-Legales")

    # Tabla de datos
    data = [["#", "Matrícula", "Nombre completo", "Plantel", "Fecha del caso", "Descripción", "Acciones adoptadas"]]

    for idx, caso in enumerate(casos, start=1):
        data.append([
            str(idx),
            caso.id_discente.matricula,
            f"{caso.id_discente.nombre} {caso.id_discente.apellido}",
            caso.id_discente.id_plantel.nombre,
            caso.fecha_caso.strftime("%d/%m/%Y"),
            caso.descripcion,
            caso.acciones_adoptadas,
        ])

    table = Table(data, colWidths=[20, 70, 120, 80, 70, 150, 150])

    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#4F81BD")),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.black),
        ('BOX', (0, 0), (-1, -1), 0.25, colors.black),
    ])

    table.setStyle(style)
    table.wrapOn(c, width, height)
    table.drawOn(c, 20, height - 80 - 20 * len(data))

    c.showPage()
    c.save()

    buffer.seek(0)
    return HttpResponse(buffer, content_type='application/pdf')